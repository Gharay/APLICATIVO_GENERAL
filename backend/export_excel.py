"""
backend/export_excel.py
Ficha Excel técnico-catastral — Agencia Nacional de Tierras
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

# ============================================================
# PALETA DE COLORES
# ============================================================
AZUL_ANT      = 'FF92D050'
AZUL_MED      = 'FFB5E6A2'
GRIS_FILA     = 'FFDAF2D0'
BLANCO        = 'FFFFFFFF'
ROJO_ALR      = 'FFFFC7CE'   # alerta presuntos baldíos
BG_PREDIO     = 'FFE2EFDA'   # verde suave — docs del predio
BG_COLINDANTE = 'FFD9E1F2'   # azul suave  — docs de colindante

# ============================================================
# UTILIDADES DE ESTILO
# ============================================================
def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

def _border():
    t = Side(style='thin', color='FF000000')
    return Border(left=t, right=t, top=t, bottom=t)

def _font_header(): return Font(bold=True, color='FF000000', size=11, name='Aptos Narrow')
def _font_label():  return Font(bold=True, size=11, name='Aptos Narrow', color='FF000000')
def _font_val():    return Font(size=11, name='Aptos Narrow')

def _aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _v(datos, clave, defecto='—'):
    if not datos:
        return defecto
    val = datos.get(clave)
    return val if val not in (None, '', 'None') else defecto

def _bool_s(valor):
    return 'Sí' if valor else 'No'

def _write(ws, fila, col, valor, font, fill, aln, border):
    c = ws.cell(row=fila, column=col,
                value=str(valor) if valor not in (None, '', 'None', '—') else '—')
    c.font = font
    c.fill = fill
    c.alignment = aln
    c.border = border
    return c

def _titulo_hoja(ws, titulo, subtitulo=''):
    ws.merge_cells('A1:B1')
    ws['A1'] = titulo
    ws['A1'].font      = Font(bold=True, size=11, color='FF000000', name='Aptos Narrow')
    ws['A1'].fill      = _fill(AZUL_ANT)
    ws['A1'].alignment = _aln('center')
    ws['A1'].border    = _border()
    ws.row_dimensions[1].height = 28
    if subtitulo:
        ws.merge_cells('A2:B2')
        ws['A2'] = subtitulo
        ws['A2'].font      = Font(italic=True, size=11, color='FF000000', name='Aptos Narrow')
        ws['A2'].alignment = _aln('center')
        return 3
    return 2

def _seccion(ws, fila, titulo, filas_datos):
    ws.merge_cells(f'A{fila}:B{fila}')
    ws[f'A{fila}'] = titulo
    ws[f'A{fila}'].font      = Font(bold=True, size=11, color='FF000000', name='Aptos Narrow')
    ws[f'A{fila}'].fill      = _fill(AZUL_MED)
    ws[f'A{fila}'].alignment = _aln('left')
    ws[f'A{fila}'].border    = _border()
    ws.row_dimensions[fila].height = 18
    fila += 1
    for i, (concepto, valor) in enumerate(filas_datos):
        bg = _fill(GRIS_FILA) if i % 2 == 0 else _fill(BLANCO)
        _write(ws, fila, 1, concepto, _font_label(), bg, _aln('left', wrap=True), _border())
        _write(ws, fila, 2, valor,    _font_val(),   bg, _aln('left', wrap=True), _border())
        ws.row_dimensions[fila].height = 15
        fila += 1
    return fila + 1

# ── Encabezado para hojas multi-columna (colindantes, documentos) ────
def _titulo_multicolumna(ws, n_cols, texto):
    ultima = get_column_letter(n_cols)
    ws.merge_cells(f'A1:{ultima}1')
    ws['A1'] = texto
    ws['A1'].font      = Font(bold=True, size=11, color='FF000000', name='Aptos Narrow')
    ws['A1'].fill      = _fill(AZUL_ANT)
    ws['A1'].alignment = _aln('center')
    ws['A1'].border    = _border()
    ws.row_dimensions[1].height = 25

# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================
def exportar_excel(datos_predio):
    wb = Workbook()
    wb.remove(wb.active)

    predio       = datos_predio.get('predio')       or {}
    localizacion = datos_predio.get('localizacion') or {}
    catastral    = datos_predio.get('catastral')    or {}
    juridica     = datos_predio.get('juridica')     or {}
    colindantes  = datos_predio.get('colindantes')  or []
    documentos   = datos_predio.get('documentos')   or []

    # ══════════════════════════════════════════════════════
    # HOJA 1 — FICHA GENERAL
    # ══════════════════════════════════════════════════════
    ws = wb.create_sheet('FICHA GENERAL')
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 45

    fila = _titulo_hoja(
        ws,
        'FICHA TÉCNICO-CATASTRAL — AGENCIA NACIONAL DE TIERRAS',
        f"Predio: {_v(predio,'id')} | FMI: {_v(predio,'fmi')} | "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    # A. Identificación
    fila = _seccion(ws, fila, 'A. IDENTIFICACIÓN DEL PREDIO', [
        ('ID del Predio',                       _v(predio, 'id')),
        ('FMI',                                  _v(predio, 'fmi')),
        ('Nombre del predio (VUR)',              _v(predio, 'nombre_predio_vur')),
        ('Propietario (VUR)',                    _v(predio, 'propietario_vur')),
        ('# Documento del Propietario (VUR)',    _v(predio, 'numero_documento_propietario_vur')),
        ('Área Poligono GC',                  _v(predio, 'area_vur', 'N/A')),
        ('¿Nace de un predio matriz?',           _bool_s(predio.get('es_derivado'))),
        ('ID Predio Matriz',                     _v(predio, 'id_predio_matriz', 'N/A')),
        ('¿Tiene predios derivados?',            _bool_s(predio.get('tiene_derivadas'))),
        ('IDs Predios Derivados',                _v(predio, 'derivadas_ids', 'N/A')),
        ('Proceso englobe / desenglobe',         _bool_s(predio.get('proceso_englobe_desenglobe'))),
        ('Estado',                               _v(predio, 'estado')),
        ('Fecha de creación',                    _v(predio, 'fecha_creacion')),
    ])

    # B. Localización
    fila = _seccion(ws, fila, 'B. LOCALIZACIÓN', [
        ('Departamento (ubicación espacial)',    _v(localizacion, 'dpto_espacial')),
        ('Municipio (ubicación espacial)',       _v(localizacion, 'municipio_espacial')),
        ('Vereda / Corregimiento (espacial)',    _v(localizacion, 'vereda_espacial')),
        ('¿Municipio diferente en VUR?',         _bool_s(localizacion.get('tiene_municipio_diferente_vur'))),
        ('Departamento (VUR)',                   _v(localizacion, 'dpto_vur', 'N/A')),
        ('Municipio (VUR)',                      _v(localizacion, 'municipio_vur', 'N/A')),
        ('Vereda / Corregimiento (VUR)',         _v(localizacion, 'vereda_vur', 'N/A')),
    ])

    # C. Catastral
    fila = _seccion(ws, fila, 'C. INFORMACIÓN CATASTRAL', [
        ('Número predial',                       _v(catastral, 'numero_predial')),
        ('FMI en R1/R2',                         _v(catastral, 'fmi_r1_r2')),
        ('Nombre del predio en R1/R2',           _v(catastral, 'nombre_predio_r1_r2')),
        ('Propietario (R1/R2)',                  _v(catastral, 'propietario')),
        ('Cédula del propietario',               _v(catastral, 'cedula_propietario')),
        ('Área de terreno — polígono (ha)',      _v(catastral, 'area_terreno_ha', 'N/A')),
        ('Destino económico',                    _v(catastral, 'destino_economico')),
        ('Referencia catastral',                 _v(catastral, 'referencia_catastral')),
    ])

    # D. Jurídica
    fila = _seccion(ws, fila, 'D. INFORMACIÓN JURÍDICA', [
        ('Nombre del predio actual',             _v(juridica, 'nombre_predio_actual')),
        ('Nombre del predio en doc. de origen',  _v(juridica, 'nombre_predio_origen')),
        ('Documento jurídico de origen',         _v(juridica, 'documento_origen_inmueble')),
        ('Área jurídica (ha)',                   _v(juridica, 'area_juridica_ha', 'N/A')),
        ('Adjudicatario inicial',                _v(juridica, 'adjudicatario_inicial')),
        ('Referencia catastral (VUR)',           _v(juridica, 'referencia_catastral_vur')),
        ('Tipo de instrumento',                  _v(juridica, 'tipo_instrumento')),
        ('Fecha de instrumento',                 _v(juridica, 'fecha_instrumento')),
        ('Tipo de predio',                       _v(juridica, 'tipo_predio')),
        ('Estado del folio',                     _v(juridica, 'estado_folio')),
        ('Fecha apertura del folio',             _v(juridica, 'fecha_apertura_folio')),
        ('Documento primera anotación',          _v(juridica, 'documento_primera_anotacion')),
        ('Documento última anotación',           _v(juridica, 'documento_ultima_anotacion')),
        ('Plano',                                _v(juridica, 'plano_juridica')),
        ('Complementaciones',                    _v(juridica, 'complementaciones')),
        ('Cabida y linderos',                    _v(juridica, 'cabida_linderos')),
        ('Salvedades',                           _v(juridica, 'salvedades')),
        ('Vida jurídica',                        _v(juridica, 'vida_juridica')),
        ('CAS',                                  _v(juridica, 'cas')),
    ])

    # E. Resumen colindantes
    validados  = sum(1 for c in colindantes if (c.get('estado_validacion') or '').upper() == 'VALIDADO')
    no_verif   = sum(1 for c in colindantes if (c.get('estado_validacion') or '').upper() == 'NO_VERIFICADO')
    baldios    = sum(1 for c in colindantes if c.get('es_presunto_baldio'))
    req_campo  = sum(1 for c in colindantes if c.get('requiere_verificacion_campo'))
    n_docs_p   = len(documentos)
    n_docs_c   = sum(len(c.get('documentos_colindante') or []) for c in colindantes)

    fila = _seccion(ws, fila, 'E. RESUMEN COLINDANTES', [
        ('Total de colindantes analizados',            len(colindantes)),
        ('Colindantes validados',           validados),
        ('Colindantes no verificados',      no_verif),
        ('Requieren verificación en campo', req_campo),
        ('Presuntos baldíos',              baldios),
    ])

    # F. Resumen documentos
    fila = _seccion(ws, fila, 'F. RESUMEN DOCUMENTOS', [
        ('Documentos del predio',           n_docs_p),
        ('Documentos de colindantes',       n_docs_c),
        ('Total documentos',               n_docs_p + n_docs_c),
    ])

    # ══════════════════════════════════════════════════════
    # HOJA 2 — G. COLINDANTES
    # ══════════════════════════════════════════════════════
    ws_col = wb.create_sheet('G. COLINDANTES')

    # NOTA: todas las listas están correctamente cerradas con ]
    cols_col = [
        'N°', 'Cardinalidad', 'FMI Colindante', 'N° Predial',
        'Nombre Predio', 'Departamento', 'Municipio', 'Vereda',
        'Prop. Inicial', 'Prop. Actual VUR',
        'Tipo', 'ID ANT', 'Levant.', 'Doc. Origen', 'Plano',
        'Val. Campo', 'Estado Valid.', 'Fuente Valid.',
        'Req. Campo', 'Presunto Baldío',
        '1ª Anotación', 'Últ. Anotación', 'Observaciones',
        '¿Matriz?', 'ID Predio Matriz', '¿Tiene derivadas?', 'IDs derivadas'
    ]
    anchos_col = [
        5,  13, 18, 17,
        22, 16, 16, 16,
        22, 22,
        16, 13, 10, 22, 12,
        12, 18, 26,
        12, 15,
        26, 26, 28,
        26, 26, 28, 28
    ]

    _titulo_multicolumna(ws_col, len(cols_col), f"G. COLINDANTES — Predio {_v(predio, 'id')}")

    for c_idx, (h, ancho) in enumerate(zip(cols_col, anchos_col), start=1):
        cell = ws_col.cell(row=2, column=c_idx, value=h)
        cell.font      = _font_header()
        cell.fill      = _fill(AZUL_MED)
        cell.alignment = _aln('center')
        cell.border    = _border()
        ws_col.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws_col.row_dimensions[2].height = 18

    for r_idx, col in enumerate(colindantes, start=3):
        es_baldio = col.get('es_presunto_baldio')
        bg = _fill(ROJO_ALR) if es_baldio else (
             _fill(GRIS_FILA) if r_idx % 2 == 1 else _fill(BLANCO))

        valores = [
            col.get('numero_colindante', r_idx - 2),
            _v(col, 'cardinalidad'),
            _v(col, 'fmi_colindante'),
            _v(col, 'numero_predial_colindante'),
            _v(col, 'nombre_predio_colindante'),
            _v(col, 'departamento_colindante', ''),
            _v(col, 'municipio_colindante', ''),
            _v(col, 'vereda_colindante', ''),
            _v(col, 'propietario_inicial_colindante'),
            _v(col, 'propietario_actual_vur_colindante'),
            _v(col, 'tipo_colindante'),
            _v(col, 'id_ant'),
            _bool_s(col.get('tiene_levantamiento')),
            _v(col, 'documento_origen_colindante'),
            _v(col, 'plano_colindante'),
            _bool_s(col.get('necesario_validar_campo')),
            _v(col, 'estado_validacion', ''),
            _v(col, 'fuente_validacion', ''),
            _bool_s(col.get('requiere_verificacion_campo')),
            _bool_s(es_baldio),
            _v(col, 'colindante_primera_anotacion', ''),
            _v(col, 'colindante_ultima_anotacion', ''),
            _v(col, 'observacion_colindante', ''),
            _bool_s(col.get('es_derivado_colindante')),
            _v(col, 'id_predio_matriz_colindante', 'N/A'),
            _bool_s(col.get('tiene_derivadas_colindante')),
            _v(col, 'derivadas_ids_colindante', 'N/A'),

        ]

        for c_idx, valor in enumerate(valores, start=1):
            _write(ws_col, r_idx, c_idx, valor, _font_val(), bg,
                   _aln('left', wrap=True), _border())
        ws_col.row_dimensions[r_idx].height = 16

    # ══════════════════════════════════════════════════════
    # HOJA 3 — H. DOCUMENTOS (predio + colindantes)
    # ══════════════════════════════════════════════════════
    ws_doc = wb.create_sheet('H. DOCUMENTOS')

    cols_doc   = [
        'Origen', 'Referencia (FMI / Nombre)', 'Tipo de documento',
        'Nombre del archivo', 'Fecha de carga', 'Tamaño (KB)',
    ]
    anchos_doc = [13, 34, 24, 44, 18, 14]

    _titulo_multicolumna(ws_doc, len(cols_doc),
                         f"H. DOCUMENTOS ADJUNTOS — Predio {_v(predio, 'id')}")

    for c_idx, (h, ancho) in enumerate(zip(cols_doc, anchos_doc), start=1):
        cell = ws_doc.cell(row=2, column=c_idx, value=h)
        cell.font      = _font_header()
        cell.fill      = _fill(AZUL_MED)
        cell.alignment = _aln('center')
        cell.border    = _border()
        ws_doc.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws_doc.row_dimensions[2].height = 18

    # ── Lista unificada ──────────────────────────────────
    filas_doc = []

    for d in documentos:
        kb = round(d.get('tamaño_bytes', 0) / 1024, 1) if d.get('tamaño_bytes') else '—'
        filas_doc.append((
            'PREDIO',
            _v(predio, 'id'),
            _v(d, 'tipo_documento'),
            _v(d, 'nombre_archivo'),
            _v(d, 'fecha_carga'),
            kb,
        ))

    for col in colindantes:
        ref = ' / '.join(
            x for x in [
                _v(col, 'fmi_colindante', ''),
                _v(col, 'nombre_predio_colindante', ''),
            ] if x and x != '—'
        ) or f"Colindante {col.get('numero_colindante', '')}"

        for dc in (col.get('documentos_colindante') or []):
            kb = round(dc.get('tamaño_bytes', 0) / 1024, 1) if dc.get('tamaño_bytes') else '—'
            filas_doc.append((
                'COLINDANTE',
                ref,
                dc.get('tipo_documento', '—') or '—',
                dc.get('nombre_archivo',  '—') or '—',
                dc.get('fecha_carga',     '—') or '—',
                kb,
            ))

    for r_idx, fila_vals in enumerate(filas_doc, start=3):
        origen = fila_vals[0]
        if r_idx % 2 == 1:
            bg = _fill(BG_PREDIO) if origen == 'PREDIO' else _fill(BG_COLINDANTE)
        else:
            bg = _fill(BLANCO)
        for c_idx, valor in enumerate(fila_vals, start=1):
            _write(ws_doc, r_idx, c_idx, valor, _font_val(), bg,
                   _aln('left', wrap=True), _border())
        ws_doc.row_dimensions[r_idx].height = 15

    if filas_doc:
        pie     = len(filas_doc) + 3
        n_pred  = sum(1 for f in filas_doc if f[0] == 'PREDIO')
        n_coln  = sum(1 for f in filas_doc if f[0] == 'COLINDANTE')
        ws_doc.merge_cells(f'A{pie}:F{pie}')
        ws_doc[f'A{pie}'] = (f"Total: {len(filas_doc)} documento(s)  "
                              f"|  Del predio: {n_pred}  |  De colindantes: {n_coln}")
        ws_doc[f'A{pie}'].font      = Font(bold=True, italic=True, size=10, name='Aptos Narrow')
        ws_doc[f'A{pie}'].fill      = _fill(AZUL_MED)
        ws_doc[f'A{pie}'].alignment = _aln('left')
        ws_doc[f'A{pie}'].border    = _border()

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
